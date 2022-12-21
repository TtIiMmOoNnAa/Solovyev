import csv
from unittest import loader

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit


class DataSet:
    def __init__(self, nazvanii_f, nazvanii_vac):
        self.nazvanii_f = nazvanii_f
        self.nazvanii_vac = nazvanii_vac

    def csv_reader(self):
        with open(self.nazvanii_f, mode='r', encoding='utf-8-sig') as citaemoe:
            zarplata_nomera = dict()
            zarplata_po_gorodam = dict()
            zarplata_po_imenam = dict()
            zarplata = dict()
            schetchik_vac = 0
            nomer_v = dict()
            nazvanii_colonok = list()
            v_num = dict()
            rows = csv.reader(citaemoe)
            for numer, stroka in enumerate(rows):
                if numer == 0:
                    nazvanii_colonok = stroka
                    dlina_naz = len(stroka)
                    dlina = len(stroka)
                elif dlina_naz == dlina and '' not in stroka:
                    a = zip(nazvanii_colonok, stroka)
                    b = dict(a)
                    odna_vac = Vacancy(b)
                    if odna_vac.year not in zarplata:
                        rez1 = odna_vac.year
                        sred = odna_vac.salary_average
                        rez2 = [sred]
                        zarplata[rez1] = rez2
                    else:
                        sred = odna_vac.salary_average
                        rez1 = odna_vac.year
                        zarplata[rez1].append(sred)
                    if odna_vac.year not in v_num:
                        rez1 = odna_vac.year
                        v_num[rez1] = 1
                    else:
                        rez1 = odna_vac.year
                        rez2 = v_num[rez1] + 1
                        v_num[rez1] = rez2
                    res = self.nazvanii_vac
                    if -1 != odna_vac.name.find(res):
                        if odna_vac.year not in zarplata_po_imenam:
                            rez1 = odna_vac.year
                            sred = odna_vac.salary_average
                            rez2 = [sred]
                            zarplata_po_imenam[rez1] = rez2
                            if odna_vac.year not in nomer_v:
                                rez1 = odna_vac.year
                                nomer_v[rez1] = 1
                            else:
                                rez1 = odna_vac.year
                                rez2 = nomer_v[rez1] + 1
                                nomer_v[rez1] = rez2
                        else:
                            rez1 = odna_vac.year
                            sred = odna_vac.salary_average
                            zarplata_po_imenam[rez1].append(sred)
                            if odna_vac.year not in nomer_v:
                                rez1 = odna_vac.year
                                nomer_v[rez1] = 1
                            else:
                                rez1 = odna_vac.year
                                rez2 = nomer_v[rez1] + 1
                                nomer_v[rez1] = rez2
                    if odna_vac.area_name not in zarplata_po_gorodam:
                        rez1 = odna_vac.area_name
                        sred = odna_vac.salary_average
                        rez2 = [sred]
                        zarplata_po_gorodam[rez1] = rez2
                    else:
                        rez1 = odna_vac.area_name
                        sred = odna_vac.salary_average
                        zarplata_po_gorodam[rez1].append(sred)
                    if odna_vac.area_name not in zarplata_nomera:
                        rez1 = odna_vac.area_name
                        zarplata_nomera[rez1] = 1
                    else:
                        rez1 = odna_vac.area_name
                        rez2 = zarplata_nomera[rez1] + 1
                        zarplata_nomera[rez1] = rez2
                    schetchik_vac = schetchik_vac + 1
            if not zarplata_po_imenam:
                zarplata_po_imenam = zarplata.copy()
                rez = self.vspomogatil_1(zarplata_po_imenam)
                rez1 = dict(rez)
                zarplata_po_imenam = rez1
                nomer_v = v_num.copy()
                res = self.vspomogatel_2(nomer_v)
                res1 = dict(res)
                nomer_v = res1
            result = dict()
            for god, zap in zarplata.items():
                a = sum(zap)
                b = len(zap)
                c = a / b
                result[god] = int(c)
            result2 = dict()
            for god, zap in zarplata_po_imenam.items():
                dlin = len(zap)
                if 0 == dlin:
                    result2[god] = 0
                else:
                    a = sum(zap)
                    b = len(zap)
                    c = a / b
                    result2[god] = int(c)
            result3 = dict()
            for god, zap in zarplata_po_gorodam.items():
                a = sum(zap)
                b = len(zap)
                c = a / b
                result3[god] = int(c)
            result4 = dict()
            for god, zap in zarplata_nomera.items():
                rez = round(zap / schetchik_vac, 4)
                result4[god] = rez
            result4 = list(self.filt1(result4))
            result4.sort(key=self.sorterovka1(), reverse=True)
            result5 = result4.copy()
            result4 = dict(result4)
            result3 = list(self.filt2(result3, result4))
            result3.sort(key=self.sorterovka2(), reverse=True)
            res123 = result3[:10]
            result3 = dict(res123)
            rez1, rez2, rez3, rez4, rez5, rez6 = self.vivo_monitor(nomer_v, result, result2, result3, result5, v_num)
            return rez1, rez2, rez3, rez4, rez5, rez6

    def vivo_monitor(self, nomer_v, result, result2, result3, result5, v_num):
        s1 = str(result)
        str1 = f'Динамика уровня зарплат по годам: {s1}'
        print(str1)
        s2 = str(v_num)
        str2 = f'Динамика количества вакансий по годам: {s2}'
        print(str2)
        s3 = str(result2)
        str3 = f'Динамика уровня зарплат по годам для выбранной профессии: {s3}'
        print(str3)
        s4 = str(nomer_v)
        str4 = f'Динамика количества вакансий по годам для выбранной профессии: {s4}'
        print(str4)
        s5 = str(result3)
        str5 = f'Уровень зарплат по городам (в порядке убывания): {s5}'
        print(str5)
        rez = result5[:10]
        rez1 = dict(rez)
        s6 = str(rez1)
        str6 = f'Доля вакансий по городам (в порядке убывания): {s6}'
        print(str6)
        return result, v_num, result2, nomer_v, result3, rez1

    def sorterovka2(self):
        return lambda znac: znac[-1]

    def filt2(self, result3, result4):
        return filter(lambda a: a[0] in list(result4.keys()), [(key, value) for key, value in result3.items()])

    def sorterovka1(self):
        return lambda znac: znac[-1]

    def filt1(self, result4):
        return filter(lambda znac: znac[-1] >= 0.01, list((kluch, znacheni) for kluch, znacheni in result4.items()))

    def vspomogatel_2(self, nomer_v):
        return list((kluch, 0) for kluch, znacheni in nomer_v.items())

    def vspomogatil_1(self, zarplata_po_imenam):
        return list((kluch, list()) for kluch, znacheni in zarplata_po_imenam.items())


class Report:
    def __init__(self, nazvano_vac, req1, req2, req3, req4, req5, req6):
        self.dan = Workbook()
        self.nazvano_vac = nazvano_vac
        self.req1 = req1
        self.req2 = req2
        self.req3 = req3
        self.req4 = req4
        self.req5 = req5
        self.req6 = req6

    def sozdan_exl(self):
        dn = self.dan.active
        dn.title = 'Статистика по годам'
        dn.append(self.dobavka1())
        for god in self.req1.keys():
            dn.append(self.dobavka2(god))
        danii = [self.dobavka3()]
        dlina_colon = list()
        for stroka in danii:
            for nomer, elem in enumerate(stroka):
                dl1 = len(dlina_colon)
                if nomer < dl1:
                    dl2 = len(elem)
                    if dlina_colon[nomer] < dl2:
                        dlina_colon[nomer] = dl2
                else:
                    dl2 = len(elem)
                    dlina_colon = dlina_colon + [dl2]

        for nomer, dl_c in enumerate(dlina_colon, 1):
            rez = dl_c + 2
            dn.column_dimensions[get_column_letter(nomer)].width = rez
        danii = list()
        danii.append(self.dobavka4())
        probeg = zip(self.req5.items(), self.req6.items())
        for (gorod1, znachen1), (gorod2, znachenii2) in probeg:
            danii.append(self.dobavka5(gorod1, gorod2, znachen1, znachenii2))
        dn2 = self.dan.create_sheet('Статистика по городам')
        for stroka in danii:
            dn2.append(stroka)
        dlina_colon = list()
        for stroka in danii:
            for nomer, elem in enumerate(stroka):
                dl1 = len(dlina_colon)
                elem = str(elem)
                if nomer < dl1:
                    dl2 = len(elem)
                    if dlina_colon[nomer] < dl2:
                        dlina_colon[nomer] = dl2
                else:
                    dl2 = len(elem)
                    dlina_colon = dlina_colon + [dl2]
        for nomer, dl_c in enumerate(dlina_colon, 1):
            rez = dl_c + 2
            dn2.column_dimensions[get_column_letter(nomer)].width = rez
        shirift_tolste = Font(bold=True)
        for chislo in 'ABCDE':
            rez = chislo + '1'
            dn[rez].font = shirift_tolste
            dn2[rez].font = shirift_tolste
        for numer, i in enumerate(self.req5):
            rez = 'E' + str(numer + 2)
            dn2[rez].number_format = '0.00%'
        shirift_tonki = Side(border_style='thin', color='00000000')
        for stroka in range(len(danii)):
            for chislo in 'ABDE':
                rez = chislo + str(stroka + 1)
                dn2[rez].border = self.tolchina1(shirift_tonki)
        for stroka, i in enumerate(self.req1):
            for chislo in 'ABCDE':
                rez = chislo + str(stroka + 1)
                dn[rez].border = self.tolchina1(shirift_tonki)
        self.dan.save(filename='report.xlsx')

    def tolchina1(self, shirift_tonki):
        return Border(left=shirift_tonki, bottom=shirift_tonki, right=shirift_tonki, top=shirift_tonki)

    def dobavka5(self, gorod1, gorod2, znachen1, znachenii2):
        rez = [gorod1, znachen1, '', gorod2, znachenii2]
        return rez

    def dobavka4(self):
        return self.dobavka5('Город', 'Город', 'Уровень зарплат', 'Доля вакансий')

    def dobavka3(self):
        rez = ['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.nazvano_vac, ' Количество вакансий',
               ' Количество вакансий - ' + self.nazvano_vac]
        return rez

    def sum_elem(self):
        return [value for value in self.req6.values()]

    def sroika_method3(self, koordinata3):
        koordinata3.yaxis.set_tick_params(labelsize=6)
        koordinata3.xaxis.set_tick_params(labelsize=8)

    def numpy_method5(self):
        return np.array(list(self.req2.keys()))

    def numpy_method4(self):
        return self.numpy_method5() - 0.4

    def stroitel_method1(self, koordinata1):
        koordinata1.xaxis.set_tick_params(labelsize=8)
        koordinata1.yaxis.set_tick_params(labelsize=8)

    def numpy_method2(self):
        return np.array(list(self.req1.keys()))

    def numpy_method1(self):
        return self.numpy_method2() - 0.4

    def nachalni_koord1(self):
        koord, ((koordinata1, koordinata2), (koordinata3, koordinata4)) = plt.subplots(nrows=2, ncols=2)
        return koordinata1, koordinata2, koordinata3, koordinata4

    # def save(self, naz):
    # self.dan.save(filename=naz)

    def dobavka2(self, god):
        rez = self.dobavlen_pdf1(god)
        return rez

    def dobavka1(self):
        rez = ['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.nazvano_vac, 'Количество вакансий',
               'Количество вакансий - ' + self.nazvano_vac]
        return rez

    def sozdan_png(self):
        koordinata1, koordinata2, koordinata3, koordinata4 = self.nachalni_koord1()
        y_perv1 = koordinata1.bar(self.numpy_method1(), self.req1.values(), width=0.4)
        y_perv2 = koordinata1.bar(self.numpy_method2(), self.req3.values(), width=0.4)
        koordinata1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        # y
        koordinata1.grid(axis='y')
        # legenda
        koordinata1.legend((y_perv1[0], y_perv2[0]), ('средняя з/п', 'з/п ' + self.nazvano_vac.lower()),
                           prop={'size': 8})
        koordinata1.set_xticks(self.numpy_method2() - 0.2, list(self.req1.keys()), rotation=90)
        self.stroitel_method1(koordinata1)
        koordinata2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        y_perv1 = koordinata2.bar(self.numpy_method4(), self.req2.values(), width=0.4)
        y_perv2 = koordinata2.bar(self.numpy_method5(), self.req4.values(), width=0.4)
        koordinata2.legend((y_perv1[0], y_perv2[0]),
                           ('Количество вакансий', 'Количество вакансий\n' + self.nazvano_vac.lower()),
                           prop={'size': 8})
        koordinata2.set_xticks(self.numpy_method5() - 0.2, list(self.req2.keys()), rotation=90)
        # y
        koordinata2.grid(axis='y')
        self.stroitel_method1(koordinata2)
        koordinata3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        koordinata3.barh(
            list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.req5.keys()))]),
            list(reversed(list(self.req5.values()))), color='blue', height=0.5, align='center')
        self.sroika_method3(koordinata3)
        # x
        koordinata3.grid(axis='x')
        koordinata4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum(self.sum_elem())
        koordinata4.pie(list(self.req6.values()) + [other], labels=list(self.req6.keys()) + ['Другие'],
                        textprops={'fontsize': 6})
        plt.tight_layout()
        plt.savefig('graph.png')

    def sozdan_pdf(self):
        env = Environment(loader=FileSystemLoader('D:/pythonProject2.1.3/templates'))
        template = env.get_template('pdf_template.html')
        elems = list()
        for god in self.req1.keys():
            elems.append(self.dobavlen_pdf1(god))
        for kluch in self.req6:
            self.req6[kluch] = self.okrugleni(kluch)
        pdf_template = template.render(
            {'name': self.nazvano_vac, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
             'stats': elems, 'stats5': self.req5, 'stats6': self.req6})
        ####dly otobragn_pdf
        config = pdfkit.configuration(wkhtmltopdf=r"C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})
        ####konec
        pdfkit.from_string(pdf_template, 'report.pdf', options={"enable-local-file-access": ""})

    def okrugleni(self, kluch):
        rez = round(self.req6[kluch] * 100, 2)
        return rez

    def dobavlen_pdf1(self, year):
        rez = [year, self.req1[year], self.req3[year], self.req2[year], self.req4[year]]
        return rez


class Vacancy:
    kurs_rubl = {
        "UZS": 0.0055,
        "USD": 60.66,
        "UAH": 1.64,
        "RUR": 1,
        "KZT": 0.13,
        "KGS": 0.76,
        "GEL": 21.74,
        "EUR": 59.90,
        "BYR": 23.91,
        "AZN": 35.68
    }

    def __init__(self, odna_vac):
        self.name = odna_vac['name']
        self.salary_from = int(float(odna_vac['salary_from']))
        self.salary_to = int(float(odna_vac['salary_to']))
        self.salary_currency = odna_vac['salary_currency']
        sum_1 = (self.salary_from + self.salary_to)
        isprav = self.kurs_rubl[self.salary_currency] * sum_1
        res = isprav / 2
        self.salary_average = res
        self.area_name = odna_vac['area_name']
        self.year = int(odna_vac['published_at'][:4])


class InputConnect:

    def __init__(self):
        str1 = 'Введите название файла: '
        self.nazvanii_f = input(str1)
        str2 = 'Введите название профессии: '
        self.nazvanii_vac = input(str2)
        final_danii = DataSet(self.nazvanii_f, self.nazvanii_vac)
        r1, r2, r3, r4, r5, r6 = final_danii.csv_reader()
        rez_xml = Report(self.nazvanii_vac, r1, r2, r3, r4, r5, r6)
        rez_xml.sozdan_exl()
        # rez_xml.save('report.xlsx')
        rez_xml.sozdan_png()
        rez_xml.sozdan_pdf()


if __name__ == '__main__':
    InputConnect()
