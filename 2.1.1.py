import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class DataSet:
    def __init__(self, nazvanii_f, nazvanii_vac):
        self.nazvanii_f = nazvanii_f
        self.nazvanii_vac = nazvanii_vac

    def csv_reader(self):
        with open(self.nazvanii_f, mode='r', encoding='utf-8-sig') as citaemoe:
            zarplata_nomera = dict()
            zarplata_po_gorodam = dict()
            zarplata_po_imenam = dict()
            zarplata = dict()
            schetchik_vac = 0
            nomer_v = dict()
            nazvanii_colonok = list()
            v_num = dict()
            rows = csv.reader(citaemoe)
            for numer, stroka in enumerate(rows):
                if numer == 0:
                    nazvanii_colonok = stroka
                    dlina_naz = len(stroka)
                    dlina = len(stroka)
                elif dlina_naz == dlina and '' not in stroka:
                    a = zip(nazvanii_colonok, stroka)
                    b = dict(a)
                    odna_vac = Vacancy(b)
                    if odna_vac.year not in zarplata:
                        rez1 = odna_vac.year
                        sred = odna_vac.salary_average
                        rez2 = [sred]
                        zarplata[rez1] = rez2
                    else:
                        sred = odna_vac.salary_average
                        rez1 = odna_vac.year
                        zarplata[rez1].append(sred)
                    if odna_vac.year not in v_num:
                        rez1 = odna_vac.year
                        v_num[rez1] = 1
                    else:
                        rez1 = odna_vac.year
                        rez2 = v_num[rez1] + 1
                        v_num[rez1] = rez2
                    res = self.nazvanii_vac
                    if -1 != odna_vac.name.find(res):
                        if odna_vac.year not in zarplata_po_imenam:
                            rez1 = odna_vac.year
                            sred = odna_vac.salary_average
                            rez2 = [sred]
                            zarplata_po_imenam[rez1] = rez2
                            if odna_vac.year not in nomer_v:
                                rez1 = odna_vac.year
                                nomer_v[rez1] = 1
                            else:
                                rez1 = odna_vac.year
                                rez2 = nomer_v[rez1] + 1
                                nomer_v[rez1] = rez2
                        else:
                            rez1 = odna_vac.year
                            sred = odna_vac.salary_average
                            zarplata_po_imenam[rez1].append(sred)
                            if odna_vac.year not in nomer_v:
                                rez1 = odna_vac.year
                                nomer_v[rez1] = 1
                            else:
                                rez1 = odna_vac.year
                                rez2 = nomer_v[rez1] + 1
                                nomer_v[rez1] = rez2
                    if odna_vac.area_name not in zarplata_po_gorodam:
                        rez1 = odna_vac.area_name
                        sred = odna_vac.salary_average
                        rez2 = [sred]
                        zarplata_po_gorodam[rez1] = rez2
                    else:
                        rez1 = odna_vac.area_name
                        sred = odna_vac.salary_average
                        zarplata_po_gorodam[rez1].append(sred)
                    if odna_vac.area_name not in zarplata_nomera:
                        rez1 = odna_vac.area_name
                        zarplata_nomera[rez1] = 1
                    else:
                        rez1 = odna_vac.area_name
                        rez2 = zarplata_nomera[rez1] + 1
                        zarplata_nomera[rez1] = rez2
                    schetchik_vac = schetchik_vac + 1
            if not zarplata_po_imenam:
                zarplata_po_imenam = zarplata.copy()
                rez = self.vspomogatil_1(zarplata_po_imenam)
                rez1 = dict(rez)
                zarplata_po_imenam = rez1
                nomer_v = v_num.copy()
                res = self.vspomogatel_2(nomer_v)
                res1 = dict(res)
                nomer_v = res1
            result = dict()
            for god, zap in zarplata.items():
                a = sum(zap)
                b = len(zap)
                c = a / b
                result[god] = int(c)
            result2 = dict()
            for god, zap in zarplata_po_imenam.items():
                dlin = len(zap)
                if 0 == dlin:
                    result2[god] = 0
                else:
                    a = sum(zap)
                    b = len(zap)
                    c = a / b
                    result2[god] = int(c)
            result3 = dict()
            for god, zap in zarplata_po_gorodam.items():
                a = sum(zap)
                b = len(zap)
                c = a / b
                result3[god] = int(c)
            result4 = dict()
            for god, zap in zarplata_nomera.items():
                rez = round(zap / schetchik_vac, 4)
                result4[god] = rez
            result4 = list(self.filt1(result4))
            result4.sort(key=self.sorterovka1(), reverse=True)
            result5 = result4.copy()
            result4 = dict(result4)
            result3 = list(self.filt2(result3, result4))
            result3.sort(key=self.sorterovka2(), reverse=True)
            res123 = result3[:10]
            result3 = dict(res123)
            rez1, rez2, rez3, rez4, rez5, rez6 = self.vivo_monitor(nomer_v, result, result2, result3, result5, v_num)
            return rez1, rez2, rez3, rez4, rez5, rez6

    def vivo_monitor(self, nomer_v, result, result2, result3, result5, v_num):
        s1 = str(result)
        str1 = f'Динамика уровня зарплат по годам: {s1}'
        print(str1)
        s2 = str(v_num)
        str2 = f'Динамика количества вакансий по годам: {s2}'
        print(str2)
        s3 = str(result2)
        str3 = f'Динамика уровня зарплат по годам для выбранной профессии: {s3}'
        print(str3)
        s4 = str(nomer_v)
        str4 = f'Динамика количества вакансий по годам для выбранной профессии: {s4}'
        print(str4)
        s5 = str(result3)
        str5 = f'Уровень зарплат по городам (в порядке убывания): {s5}'
        print(str5)
        rez = result5[:10]
        rez1 = dict(rez)
        s6 = str(rez1)
        str6 = f'Доля вакансий по городам (в порядке убывания): {s6}'
        print(str6)
        return result, v_num, result2, nomer_v, result3, rez1

    def sorterovka2(self):
        return lambda znac: znac[-1]

    def filt2(self, result3, result4):
        return filter(lambda a: a[0] in list(result4.keys()), [(key, value) for key, value in result3.items()])

    def sorterovka1(self):
        return lambda znac: znac[-1]

    def filt1(self, result4):
        return filter(lambda znac: znac[-1] >= 0.01, list((kluch, znacheni) for kluch, znacheni in result4.items()))

    def vspomogatel_2(self, nomer_v):
        return list((kluch, 0) for kluch, znacheni in nomer_v.items())

    def vspomogatil_1(self, zarplata_po_imenam):
        return list((kluch, list()) for kluch, znacheni in zarplata_po_imenam.items())


class Report:
    def __init__(self, nazvanii_vac, r1, r2, r3, r4, r5, r6):
        self.dan = Workbook()
        self.nazvanii_vac = nazvanii_vac
        self.r1 = r1
        self.r2 = r2
        self.r3 = r3
        self.r4 = r4
        self.r5 = r5
        self.r6 = r6

    def sozdan_exl(self):
        znacheni1 = self.dan.active
        znacheni1.title = 'Статистика по годам'
        znacheni1.append(self.shapka_xl())
        for god in self.r1.keys():
            rez = list()
            a = self.r1[god]
            b = self.r3[god]
            c = self.r2[god]
            d = self.r4[god]
            rez = [god, a, b, c, d]
            znacheni1.append(rez)
        danii = [self.stroka_nazvan2()]
        shirina = list()
        for strok in danii:
            for nomer, danoe in enumerate(strok):
                dlina = len(shirina)
                if nomer < dlina:
                    dl = len(danoe)
                    a = shirina[nomer]
                    if a < dl:
                        shirina[nomer] = dl
                else:
                    dl = len(danoe)
                    shirina += [dl]
        for nomer, shirina_colon in enumerate(shirina, 1):
            rez = shirina_colon + 2
            self.sumat_colom(nomer, rez, znacheni1)
        danii = list()
        danii.append(self.danii_dobav())
        perebor = zip(self.r5.items(), self.r6.items())
        for (gorod1, znachenii1), (gorod2, znachenii2) in perebor:
            danii.append(self.doavlenii_danix1(gorod1, gorod2, znachenii1, znachenii2))
        rez_b = self.dan.create_sheet('Статистика по городам')
        znachenii2 = rez_b
        for strok in danii:
            # dobavleni z2
            znachenii2.append(strok)
        shirina = list()
        for strok in danii:
            for nomer, danoe in enumerate(strok):
                danoe = str(danoe)
                dlina = len(shirina)
                if nomer < dlina:
                    dl = len(danoe)
                    zna_a = shirina[nomer]
                    if zna_a < dl:
                        shirina[nomer] = dl
                else:
                    dl = len(danoe)
                    rez = shirina + [dl]
                    shirina = rez
        for nomer, shirina_colon in enumerate(shirina, 1):
            self.sumat_colom(nomer, shirina_colon + 2, znachenii2)
        sirift = Font(bold=True)
        # naxogdei prefex bukv
        for col in 'ABCDE':
            rez_1 = col + '1'
            znacheni1[rez_1].font = sirift
            znachenii2[rez_1].font = sirift
        pr = self.r5
        for numer, i in enumerate(pr):
            rez = 'E' + str(numer + 2)
            znachenii2[rez].number_format = '0.00%'
        sirift_tonk = Side(border_style='thin', color='00000000')
        dlinaa = len(danii)
        for strok in range(dlinaa):
            # obxod
            for col in 'ABDE':
                rez_2 = col + str(strok + 1)
                rez = self.sirina_border(sirift_tonk)
                znachenii2[rez_2].border = rez
        self.r1[1] = 1
        for strok, i in enumerate(self.r1):
            # obxod
            for col in 'ABCDE':
                rez_2 = col + str(strok + 1)
                rez = self.sirina_border(sirift_tonk)
                znacheni1[rez_2].border = rez
        self.dan.save('report.xlsx')

    def sirina_border(self, sirift_tonk):
        return Border(left=sirift_tonk, bottom=sirift_tonk, right=sirift_tonk, top=sirift_tonk)

    def doavlenii_danix1(self, gorod1, gorod2, znachenii1, znachenii2):
        return [gorod1, znachenii1, '', gorod2, znachenii2]

    def danii_dobav(self):
        return self.doavlenii_danix1('Город', 'Город', 'Уровень зарплат', 'Доля вакансий')

    def sumat_colom(self, nomer, rez, znacheni1):
        znacheni1.column_dimensions[get_column_letter(nomer)].width = rez

    def stroka_nazvan2(self):
        r_a = self.nazvanii_vac
        return ['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + r_a, ' Количество вакансий',
                ' Количество вакансий - ' + r_a]

    def shapka_xl(self):
        r_a = self.nazvanii_vac
        return ['Год', 'Средняя зарплата', 'Средняя зарплата - ' + r_a, 'Количество вакансий',
                'Количество вакансий - ' + r_a]


class Vacancy:
    kurs_rubl = {
        "UZS": 0.0055,
        "USD": 60.66,
        "UAH": 1.64,
        "RUR": 1,
        "KZT": 0.13,
        "KGS": 0.76,
        "GEL": 21.74,
        "EUR": 59.90,
        "BYR": 23.91,
        "AZN": 35.68
    }

    def __init__(self, odna_vac):
        self.name = odna_vac['name']
        self.salary_from = int(float(odna_vac['salary_from']))
        self.salary_to = int(float(odna_vac['salary_to']))
        self.salary_currency = odna_vac['salary_currency']
        sum_1 = (self.salary_from + self.salary_to)
        isprav = self.kurs_rubl[self.salary_currency] * sum_1
        res = isprav / 2
        self.salary_average = res
        self.area_name = odna_vac['area_name']
        self.year = int(odna_vac['published_at'][:4])


class InputConnect:

    def __init__(self):
        str1 = 'Введите название файла: '
        self.nazvanii_f = input(str1)
        str2 = 'Введите название профессии: '
        self.nazvanii_vac = input(str2)
        final_danii = DataSet(self.nazvanii_f, self.nazvanii_vac)
        r1, r2, r3, r4, r5, r6 = final_danii.csv_reader()
        rez_xml = Report(self.nazvanii_vac, r1, r2, r3, r4, r5, r6)
        rez_xml.sozdan_exl()


if __name__ == '__main__':
    InputConnect()

