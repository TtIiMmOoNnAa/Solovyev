import csv
import re
from datetime import datetime
import os
from statistics import mean
import openpyxl
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import ticker
import pdfkit
from jinja2 import Environment, FileSystemLoader
import doctest



class DataSet:
    """Класс для чтения и обработки данных данного файла с расширением csv.
    Attributes:
    vacancies (list[]): Список всех строк данного файла
    shapka (list[]): Список с названиями заголовков данного файла
    obrabot_vse_vac (list[]): Отформатированный список (от тегов и т.п.) всех строк данного файла
    """

    def __init__(self, file_name):
        """Инициализирует объект DataSet, выполняет чтение и обработку данных csv файла
        Args:
        file_name (str): Название файла
        """
        if 0 == os.stat(file_name).st_size:
            rez = "Пустой файл"
            vixod_s_ochibkoi(rez)

        self.vacancies = [stroka for stroka in csv.reader(open(file_name, encoding="utf_8_sig"))]
        self.shapka = self.vacancies[0]
        self.obrabot_vse_vac = [stroka for stroka in self.vacancies[1:] if
                                len(stroka) == len(self.shapka) and stroka.count('') == 0]
        dlin = len(self.obrabot_vse_vac)

        if 0 == dlin:
            rez = 'Нет данных'
            vixod_s_ochibkoi(rez)


def vixod_s_ochibkoi(soobchenii):
    """Осуществляет принудительный выход из программы с сообщением о проблеме.
    Args:
    soobchenii (str): Сообщение, которое выводится при проблеме
    """
    print(soobchenii)
    exit(0)


class PodgotovkaDanix:
    """Класс для подготовки данных вакансии.
        """

    def __init__(self, shapka, obrabotan_vac, vviden_dani):
        """Инициализирует класс PodgotovkaDanix
        Args:
        shapka (list): список с названиями заголовков столбцов
        obrabotan_vac (list): список всех отформатированных строк данного файла
        vviden_dani Input): экземпляр класса Input
        """
        zarplat = StatistickaZarplat()
        num = StatistickaKolichestva()
        rez = FinalStatisticka()

        for odna_vac in obrabotan_vac:
            delen = map(self.ochistcka_strok, odna_vac)
            r_delen = zip(shapka, delen)
            dic_r = dict(r_delen)
            obrabotan_data = Vacancy(dic_r)
            obrabotan_data.sdelat_datty(zarplat, num, vviden_dani.vacancy_name)
        rez.vivod_itog_dan(vviden_dani.vacancy_name, zarplat, num)
        report = Report(vviden_dani.vacancy_name, rez.list_by_year, rez.list_by_area, rez.others)

    @staticmethod
    def ochistcka_strok(stroka):
        """Удаляет теги html из строки
        Args:
        stroka (str): строка из которой необходимо удалить html код
        Returns:
        list[str] or list: строка из которой удалили html теги код
        >>> PodgotovkaDanix.ochistcka_strok("Привет <body>gdsgsd</body> <a>gaffasf</a>")
        'Привет gdsgsd gaffasf'
        """
        stroka = re.sub('<.*?>', '', stroka)
        stroka = stroka.replace("\r\n", "\n")
        result_d = [' '.join(word.split()) for word in stroka.split('\n')]
        dl = len(result_d)
        if 1 == dl:
            return result_d[0]
        else:
            return result_d


perevod_valut = {
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "UZS": "Узбекский сум",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
}

kurs_k_rubl = {
    "Манаты": 35.68,
    "Белорусские рубли": 23.91,
    "Евро": 59.90,
    "Грузинский лари": 21.74,
    "Киргизский сом": 0.76,
    "Тенге": 0.13,
    "Рубли": 1,
    "Гривны": 1.64,
    "Доллары": 60.66,
    "Узбекский сум": 0.0055,
}


class Report:
    """Класс для создания отчета.
    Attributes:
    stroki_po_godam (list[]): Список названий колонок первой страницы пдф файла
    stroki_po_gorodam (list[]): Список названий колонок второй страницы пдф файла
    kniga (Workbook): Экземпляр класса Workbook
    kartinca_po_godam (Worksheet): Экземпляр класса Worksheet
    kartinca_po_gorodam (Worksheet): Экземпляр класса Worksheet
    """

    stroki_po_gorodam = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
    stroki_po_godam = ["Год", "Средняя зарплата", "Средняя зарплата - ", "Количество вакансий",
                       "Количество вакансий - "]

    def __init__(self, nazvan_vac, stroki_po_godam, stroki_po_gorodam, doly):
        """Инициализирует объект Report, создаёт экземпляр Workbook и Worksheet
        Args:
        nazvan_vac (str): Название выбранной пользователям профессии
        stroki_po_godam (list[]): Список словарей распределенных по годам
        stroki_po_gorodam (list[]): Список словарей распределенных по городам
        doly (int): Доля вакансий, не вошедших в топ-10 по количеству
        """
        self.kniga = openpyxl.Workbook()
        self.kniga.remove(self.kniga.active)
        self.kartinca_po_godam = self.kniga.create_sheet("Статистика по годам")
        self.kartinca_po_gorodam = self.kniga.create_sheet("Статистика по городам")
        self.sozdani_xml(nazvan_vac, stroki_po_godam, stroki_po_gorodam)
        self.sozdan_kartinok(nazvan_vac, stroki_po_godam, stroki_po_gorodam, doly)
        self.sozdani_pdf(nazvan_vac, stroki_po_godam, stroki_po_gorodam)

    def sozdani_xml(self, naz_vac, spisok_po_godam, spisok_po_gorodam):
        """Генерирует файл excel (таблицу) с необходимой статистикой по заданию
        Args:
        naz_vac (str): Название выбранной пользавателям профессии
        spisok_po_godam (list[]): Список словарей распределенных по годам
        spisok_po_gorodam (list[]): Список словарей распределенных по городам
        """
        granic = Side(border_style="thin", color="000000")
        self.znach_dly_table(naz_vac, spisok_po_godam, spisok_po_gorodam, granic)
        self.shirina(granic)
        self.kniga.save("report.xlsx")

    @staticmethod
    def sozdan_kartinok(naz_vac, spisok_po_godam, spisok_po_gorodam, doly):
        """Генерирует файл .png (изображение) с необходимой статистикой
        Args:
        naz_vac (str): Название выбранной пользователям профессии
        spisok_po_godam (list[]): Список словарей распределенных по годам
        spisok_po_gorodam (list[]): Список словарей распределенных по городам
        doly (int): Доля вакансий, не вошедших в топ-10 по количеству их
        """
        dlin = len(spisok_po_godam[0].keys())

        znach_po_x = np.arange(dlin)
        shirina = 0.4
        rez1 = znach_po_x - shirina / 2
        x_1 = rez1
        rez2 = znach_po_x + shirina / 2
        x_2 = rez2
        grafic = plt.figure()
        a = 221

        znach_po_x_it = grafic.add_subplot(a)
        znach_po_x_it.set_title("Уровень зарплат по годам")
        r1 = spisok_po_godam[0].values()
        znach_po_x_it.bar(x_1, r1, shirina, label="средняя з/п")
        r2 = spisok_po_godam[1].values()
        znach_po_x_it.bar(x_2, r2, shirina, label=f"з/п {naz_vac}")
        znach_po_x_it.legend(fontsize=8)
        r3 = spisok_po_godam[0].keys()
        znach_po_x_it.set_xticks(znach_po_x, r3, rotation="vertical")
        znach_po_x_it.tick_params(axis="both", labelsize=8)
        znach_po_x_it.grid(True, axis="y")
        b = 222

        znach_po_x_it = grafic.add_subplot(b)
        znach_po_x_it.set_title("Количество вакансий по годам")
        r4 = spisok_po_godam[2].values()
        znach_po_x_it.bar(x_1, r4, shirina, label="Количество вакансий")
        r5 = spisok_po_godam[3].values()
        znach_po_x_it.bar(x_2, r5, shirina, label=f"Количество вакансий \n{naz_vac}")
        znach_po_x_it.tick_params(axis="both", labelsize=8)
        r6 = spisok_po_godam[2].keys()
        znach_po_x_it.set_xticks(znach_po_x, r6, rotation="vertical")
        znach_po_x_it.legend(fontsize=8)
        znach_po_x_it.grid(True, axis="y")

        shirina = 0.8
        dlx = len(spisok_po_gorodam[0].keys())
        y_ticks_cities = np.arange(dlx)
        y_ticks_cities_names = dict()
        d = spisok_po_gorodam[0].items()
        for kluch, znachen in d:
            if " " in kluch or "-" in kluch:
                kluch = kluch.replace("-", "-\n")
                kluch = kluch.replace(" ", "\n")
            y_ticks_cities_names[kluch] = znachen
        c = 223

        znach_po_x_it = grafic.add_subplot(c)
        s1 = "Уровень зарплат по городам"
        znach_po_x_it.set_title(s1)
        r7 = spisok_po_gorodam[0].values()
        znach_po_x_it.barh(y_ticks_cities, r7, shirina, align="center")
        cc = 40000
        znach_po_x_it.xaxis.set_major_locator(ticker.MultipleLocator(cc))
        znach_po_x_it.set_yticks(y_ticks_cities, labels=y_ticks_cities_names.keys(), horizontalalignment="right",
                                 verticalalignment="center")
        znach_po_x_it.tick_params(axis="x", labelsize=8)
        znach_po_x_it.tick_params(axis="y", labelsize=6)
        znach_po_x_it.invert_yaxis()
        znach_po_x_it.grid(True, axis="x")

        znach_po_x_it = grafic.add_subplot(224)
        znach_po_x_it.set_title("Доля вакансий по городам")
        spisok_po_gorodam[1]["Другие"] = doly
        znach_po_x_it.pie(spisok_po_gorodam[1].values(), labels=spisok_po_gorodam[1].keys(), textprops={'size': 6})
        znach_po_x_it.axis('equal')

        plt.tight_layout()
        plt.savefig("graph.png")

    @staticmethod
    def sozdani_pdf(nazvan_vac, slovar_po_godam, slovar_po_gorodam):
        """Генерирует файл .pdf с необходимой статистикой
        Args:
        nazvan_vac (str): Название выбранной пользователям профессии
        slovar_po_godam (list[]): Список словарей распределенных по годам
        slovar_po_gorodam (list[]): Список словарей распределенных по городам
        """
        env = Environment(loader=FileSystemLoader("D:/pythonProject2.1.2/templates"))
        template = env.get_template('pdf_template.html')
        pdf_template = template.render(
            {'name': nazvan_vac, 'list_by_year': slovar_po_godam, 'list_by_area': slovar_po_gorodam,
             'area_td_1': list(slovar_po_gorodam[0].keys()), 'area_td_2': list(slovar_po_gorodam[0].values()),
             'area_td_3': list(slovar_po_gorodam[1].keys()), 'area_td_4': list(slovar_po_gorodam[1].values())})
        options = {'enable-local-file-access': None}
        config = pdfkit.configuration(wkhtmltopdf=r"C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def shirina(self, thins):
        """Устанавливает необходимую ширину столбцов страниц, устанавливает необходимую толщину границ ячеек,
        а также необходимый формат для ячейки с процентами
        Args:
        thins (Side): Экземпляр класса Side, задающий ширину границы ячейки
        """
        diam = dict()
        for stroka in self.kartinca_po_godam.rows:
            for elem in stroka:
                if elem.value:
                    dlin1 = len(str(elem.value))
                    rez = max((diam.get(elem.column_letter, 0), dlin1))
                    diam[elem.column_letter] = rez
        progon = diam.items()
        for row, znach in progon:
            self.kartinca_po_godam.column_dimensions[row].width = znach + 2

        diam = dict()
        for stroka in self.kartinca_po_gorodam.rows:
            for elem in stroka:
                if elem.value:
                    dlin1 = len(str(elem.value))
                    rez = max((diam.get(elem.column_letter, 0), dlin1))
                    diam[elem.column_letter] = rez
        progon2 = diam.items()
        for row, znach in progon2:
            self.kartinca_po_gorodam.column_dimensions[row].width = znach + 2

        # Изменение ячеек страниц
        b = 17
        for num in range(b):
            c = 5
            for mun in range(c):
                self.kartinca_po_godam.cell(row=num + 1, column=mun + 1).border = Border(top=thins, bottom=thins,
                                                                                         left=thins,
                                                                                         right=thins)
        d = 11
        for num in range(d):
            e = 5
            for mun in range(e):
                if 2 != mun:
                    self.kartinca_po_gorodam.cell(row=num + 1, column=mun + 1).border = Border(top=thins, bottom=thins,
                                                                                               left=thins,
                                                                                               right=thins)
        f = 10
        for num in range(f):
            rez = "0.00%"
            self.kartinca_po_gorodam.cell(row=num + 2, column=5).number_format = rez
        g = 5

        for num in range(g):
            self.kartinca_po_godam.cell(row=1, column=num + 1).font = Font(bold=True)
            self.kartinca_po_gorodam.cell(row=1, column=num + 1).font = Font(bold=True)

    def znach_dly_table(self, naz_vac, slovar_po_godam, slovar_po_gorodam, thins):
        """Устанавливает необходимые значение для таблицы
        Args:
        naz_vac (str): Название выбранной пользовыателям профессии
        slovar_po_godam (list[]): Список словарей распределенных по годам
        slovar_po_gorodam (list[]): Список словарей распределенных по городам
        thins (Side): Экземпляр класса Side задающий ширину границы ячейки
        """
        # Для первого листа
        for numer, znach in enumerate(self.stroki_po_godam, 1):
            self.kartinca_po_godam.cell(row=1, column=numer).value = znach + naz_vac if " - " in znach else znach
        a = slovar_po_godam[0].items()

        for god, znach in a:
            a1 = slovar_po_godam[1][god]
            a2 = slovar_po_godam[2][god]
            a3 = slovar_po_godam[3][god]
            self.kartinca_po_godam.append([god, znach, a1, a2, a3])

        # Для второго листа
        for numer, znach in enumerate(self.stroki_po_gorodam, 1):
            self.kartinca_po_gorodam.cell(row=1, column=numer).value = znach
        dlin4 = len(slovar_po_gorodam[0])

        for numer in range(dlin4):
            a1 = list(slovar_po_gorodam[0].keys())[numer]
            a2 = list(slovar_po_gorodam[0].values())[numer]
            a3 = list(slovar_po_gorodam[1].keys())[numer]
            a4 = list(slovar_po_gorodam[1].values())[numer]
            self.kartinca_po_gorodam.append([a1, a2, "", a3, a4])


class Vacancy:
    """Класс для представления одной вакансии.
    Attributes:
    name (str): Название вакансии
    salary_from (int or float): Нижняя граница вилки оклада
    salary_to (int or float): Верхняя граница вилки оклада
    salary_currency (str): Валюта оклада
    area_name (str): Город, в котором предоставляется вакансия
    published_at (str): Дата публикации вакансии
    salary (int): Средняя зарплата в рублях
    """

    def __init__(self, odna_vacanc):
        """Инициализирует объект Vacancy и форматирует его аттрибуты, подсчитывает среднюю зарплату в рублях
        Args:
        odna_vacanc (dict): Словарь, содержащий все данные об одной вакансии
        """
        rea = odna_vacanc.items()
        for kluch, znachen in rea:
            if 'salary_currency' == kluch:
                self.salary_currency = perevod_valut[znachen]
            elif "salary_to" == kluch:
                self.salary_to = '{:,}'.format(int(float(znachen))).replace(',', ' ')
            elif "salary_from" == kluch:
                self.salary_from = '{:,}'.format(int(float(znachen))).replace(',', ' ')
            elif 'published_at' == kluch:
                # data vremy
                self.published_at = int(datetime.strptime(znachen, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y"))
            elif "name" == kluch:
                self.name = znachen
            elif "area_name" == kluch:
                self.area_name = znachen
        a = float(self.salary_from.replace(' ', '')) + float(self.salary_to.replace(' ', ''))
        b = int(kurs_k_rubl[self.salary_currency] * a)
        self.salary = b // 2

    # sdelat_datty

    def sdelat_datty(self, slovar_po_zar, slovar_po_num, vac):
        """Заполняет словари класса PodgotovkaDanix значениями
        Args:
        slovar_po_zar (StaticSalary): Экземпляр класса StaticSalary
        slovar_po_num (StaticCount): экземпляр класса StaticCount
        vac (str): Название профессии
        """
        god = self.published_at
        komp = self.area_name
        rez = slovar_po_num.vac_count[god] + 1

        slovar_po_num.vac_count[god] = rez
        slovar_po_zar.salaries[god] += [self.salary]
        a1 = slovar_po_num.vacancies_areas.keys()

        if komp in a1:
            slovar_po_num.vacancies_areas[komp] += 1
        else:
            slovar_po_num.vacancies_areas[komp] = 1
        a2 = slovar_po_zar.salaries_areas.keys()

        if komp in a2:
            slovar_po_zar.salaries_areas[komp] += [self.salary]
        else:
            slovar_po_zar.salaries_areas[komp] = [self.salary]

        if vac != "" and vac in self.name:
            slovar_po_num.vacancy_count[god] += 1
            slovar_po_zar.vacancy_salaries[god] += [self.salary]
        rez1 = slovar_po_zar.vacancies + 1

        slovar_po_zar.vacancies = rez1


class StatistickaZarplat:
    """Класс для получения статистики по зарплатам из данного файла.
    Attributes:
    salaries (dict): Динамика уровня зарплат по годам данным
    vacancy_salaries (dict): Динамика уровня зарплат по годам для выбранной профессии
    salaries_areas (dict): Уровень зарплат по городам (в порядке убывания) - только первые топ - 10 значений
    vacancies (int): Количество всех вакансий в общем
    >>> StatistickaZarplat().salaries
    {2007: [], 2008: [], 2009: [], 2010: [], 2011: [], 2012: [], 2013: [], 2014: [], 2015: [], 2016: [], 2017: [], 2018: [], 2019: [], 2020: [], 2021: [], 2022: []}
    >>> StatistickaZarplat().vacancy_salaries
    {2007: [], 2008: [], 2009: [], 2010: [], 2011: [], 2012: [], 2013: [], 2014: [], 2015: [], 2016: [], 2017: [], 2018: [], 2019: [], 2020: [], 2021: [], 2022: []}
    """

    @staticmethod
    def proverka_dlin_slovar(d, v=""):
        """Осуществляет проверку длины получаемого словаря
        Args:
        d (dict): Словарь, который нужно проверить
        v (str): Название введенной профессии
        Returns:
        (dict): Отформатированный словарь
        >>> StatistickaZarplat().proverka_dlin_slovar({}, "Аналитик")
        {2022: 0}
        >>> StatistickaZarplat().proverka_dlin_slovar({"21": [10, 12, 15], "423": [26, 25, 49], "54": [35, 39, 34]}, "Аналитик")
        {'21': 12, '423': 33, '54': 36}
        """
        now_slovar = dict()
        a1 = d.items()
        for nazvan, znacheni in a1:
            dl_sl = len(znacheni)
            if 0 != dl_sl:
                rez = int(mean(znacheni))
                now_slovar[nazvan] = rez
        dl3 = len(now_slovar)
        if 0 == dl3 and v != '':
            return {2022: 0}
        return now_slovar

    salaries: dict = dict()
    vacancy_salaries: dict = dict()
    salaries_areas: dict = dict()
    vacancies: int = 0

    def __init__(self):
        """Инициализирует объект StatistickaZarplat, подготавливает словари для дальнейшей записи
        """
        a, b = 2007, 2023
        for numers in range(a, b):
            self.salaries[numers] = list()
            self.vacancy_salaries[numers] = list()


class StatistickaKolichestva:
    """Класс для получения статистики по количеству вакансий из данного файла.
    Attributes:
    vac_count (dict): Динамика количества вакансий по годам
    vacancy_count (dict): Динамика количества вакансий по годам для выбранной профессии
    vacancies_areas (dict): Доля вакансий по городам (в порядке убывания) - только первые топ - 10 значени
    >>> StatistickaKolichestva().vacancy_count
    {2007: 0, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}
    """
    vac_count: dict = dict()
    vacancy_count: dict = dict()
    vacancies_areas: dict = dict()

    @staticmethod
    def proverka_znach_slovar(d, vac=""):
        """Осуществляет проверку значения полученного словаря
        Args:
        d (dict): Словарь, который нужно проверить
        vac (str): Название введенной профессии
        Returns:
        (dict): Отформатированный словарь
        >>> StatistickaKolichestva().proverka_znach_slovar({}, "Аналитик")
        {2022: 0}
        >>> StatistickaKolichestva().proverka_znach_slovar(kurs_k_rubl, "Аналитик")
        {'Манаты': 35.68, 'Белорусские рубли': 23.91, 'Евро': 59.9, 'Грузинский лари': 21.74, 'Киргизский сом': 0.76, 'Тенге': 0.13, 'Рубли': 1, 'Гривны': 1.64, 'Доллары': 60.66, 'Узбекский сум': 0.0055}
        """
        now_slovar = dict()
        a = d.items()
        for nazvan, znach in a:
            if 0 != znach:
                now_slovar[nazvan] = znach
        dlin = len(now_slovar)
        if 0 == dlin and vac != "":
            return {2022: 0}
        return now_slovar

    def __init__(self):
        """Инициализирует объект StatistickaKolichestva, подготавливает словари для дальнейшей записи
        """
        a, b = 2007, 2023
        for numers in range(a, b):
            self.vac_count[numers] = 0
            self.vacancy_count[numers] = 0


class FinalStatisticka:
    """Класс для получения результата обработки статистики итоговой.
    Attributes:
    list_by_year (list): Список словарей по годам
    list_by_area (list): Список словарей по городам
    others (int): Доля вакансий, который не попали в 10
    """

    @staticmethod
    def sortirovk_slovar(slovar):
        """Сортирует словарь dictionary по значениям
        Args:
        slovar (dict): Словарь, который нужно отсортировать по значениям
        Returns:
        otsorterovan_slovar (dict): Отсортированный по значениям словарь готовый
        """
        a = slovar.items()
        sorted_tuples = sorted(a, key=lambda item: item[1], reverse=True)[:10]
        otsorterovan_slovar = {kluch: znach for kluch, znach in sorted_tuples}
        return otsorterovan_slovar

    list_by_year: list
    list_by_area: list
    others: int

    def vichisl_procenta(self, zarplat, kolichestvo):
        """Вычисляет процент вакансий из общего количества в словарях по городам и оставляет те значения,
        процент вакансий которых больше или равен 1.
        Attributes:
        zarplat (StatistickaZarplat): экземпляр класса StatistickaZarplat
        kolichestvo (StatistickaKolichestva): экземпляр класса StatistickaKolichestva
        """
        vac_po_gorodam = dict()
        zarplat_po_gor = dict()
        self.others = 0
        r1 = kolichestvo.vacancies_areas.items()
        for kluch, znach in r1:
            rez = znach / zarplat.vacancies
            percent = rez
            axe = 0.01
            if axe <= percent:
                res1 = mean(zarplat.salaries_areas[kluch])
                zarplat_po_gor[kluch] = int(res1)
                res2 = round(percent, 4)
                vac_po_gorodam[kluch] = res2
            else:
                self.others += percent
        zarplat.salaries_areas = zarplat_po_gor
        rr = vac_po_gorodam
        kolichestvo.vacancies_areas = rr

    def vivod_itog_dan(self, vac, zarplata, kolichest):
        """Печатает на экран все словари и формирует списки со словарями по годам и городам итоговые
        Args:
        vac (str): Название введенной профессии
        zarplata (StatistickaZarplat): экземпляр класса StatistickaZarplat
        kolichest (StatistickaKolichestva): экземпляр класса StatistickaKolichestva
        """
        self.vichisl_procenta(zarplata, kolichest)
        r1 = zarplata.proverka_dlin_slovar(zarplata.salaries)
        viv_1 = r1
        r2 = kolichest.proverka_znach_slovar(kolichest.vac_count)
        viv_2 = r2
        r3 = zarplata.proverka_dlin_slovar(zarplata.vacancy_salaries, vac)
        viv_3 = r3
        r4 = kolichest.proverka_znach_slovar(kolichest.vacancy_count, vac)
        viv_4 = r4
        viv_5 = self.sortirovk_slovar(zarplata.salaries_areas)
        viv_6 = self.sortirovk_slovar(kolichest.vacancies_areas)
        a1 = "Динамика уровня зарплат по годам:"

        print(a1, viv_1)
        a2 = "Динамика количества вакансий по годам:"
        print(a2, viv_2)
        a3 = "Динамика уровня зарплат по годам для выбранной профессии:"
        print(a3, viv_3)
        a4 = "Динамика количества вакансий по годам для выбранной профессии:"
        print(a4, viv_4)
        a5 = "Уровень зарплат по городам (в порядке убывания):"
        print(a5, viv_5)
        a6 = "Доля вакансий по городам (в порядке убывания):"
        print(a6, viv_6)
        self.list_by_year = [viv_1, viv_3, viv_2, viv_4]
        self.list_by_area = [viv_5, viv_6]


class Input:
    """Класс для принятия пользовательского ввода.
    Attributes:
    file_name (string): Название файла
    vacancy_name (string): Название профессии
    """

    def __init__(self):
        """Инициализирует объект Input, принимает пользовательский ввод данных"""
        a = 'Введите название файла: '
        self.file_name = input(a)
        b = 'Введите название профессии: '
        self.vacancy_name = input(b)


glav_dani = Input()
dataset = DataSet(glav_dani.file_name)
shapka = dataset.shapka
obrabotan_vacans = dataset.obrabot_vse_vac
itog_vac = PodgotovkaDanix(shapka, obrabotan_vacans, glav_dani)
